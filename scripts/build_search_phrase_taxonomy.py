import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "edited_home_project_search_phrases.xlsx"
OUTPUT_DIR = ROOT / "outputs" / "search_phrase_taxonomy"
TAXONOMY_PATH = ROOT / "backend" / "app" / "search_phrase_taxonomy.json"
WORKBOOK_DATA_PATH = OUTPUT_DIR / "workbook_data.json"


CANONICAL_BY_CLUSTER = {
    "Getting Rid of Mice": {
        "id": "get-rid-of-mice",
        "canonical_query": "how to get rid of mice",
        "title": "Get Rid of Mice in a House",
        "category": "pest_control",
        "safety_level": "standard",
    },
    "Replacing an Air Filter": {
        "id": "replace-hvac-air-filter",
        "canonical_query": "how to replace an HVAC air filter",
        "title": "Replace an HVAC Air Filter",
        "category": "hvac",
        "safety_level": "low",
    },
    "Shingling a Roof": {
        "id": "shingle-roof",
        "canonical_query": "how to shingle a roof",
        "title": "Shingle an Asphalt Roof",
        "category": "roofing_gutters",
        "safety_level": "high",
    },
    "Installing Solar Panels": {
        "id": "install-solar-panels",
        "canonical_query": "how to install solar panels",
        "title": "Install Solar Panels",
        "category": "electrical_energy",
        "safety_level": "licensed_or_high_risk",
    },
    "Installing a Heat Pump": {
        "id": "install-heat-pump",
        "canonical_query": "how to install a heat pump",
        "title": "Install a Heat Pump",
        "category": "hvac",
        "safety_level": "licensed_or_high_risk",
    },
    "Replacing a Light Switch": {
        "id": "replace-light-switch",
        "canonical_query": "how to replace a light switch",
        "title": "Replace a Light Switch",
        "category": "electrical",
        "safety_level": "electrical",
    },
    "Adding a GFCI Outlet": {
        "id": "install-gfci-outlet",
        "canonical_query": "how to install a GFCI outlet",
        "title": "Install a GFCI Outlet",
        "category": "electrical",
        "safety_level": "electrical",
    },
    "Sink Disposal Repair": {
        "id": "repair-garbage-disposal",
        "canonical_query": "how to repair a garbage disposal",
        "title": "Repair a Sink Garbage Disposal",
        "category": "plumbing",
        "safety_level": "electrical_plumbing",
    },
    "Tiling a Shower": {
        "id": "tile-shower",
        "canonical_query": "how to tile a shower",
        "title": "Tile a Shower",
        "category": "flooring_tile",
        "safety_level": "standard",
    },
    "Using Spray Foam": {
        "id": "apply-spray-foam-insulation",
        "canonical_query": "how to use spray foam insulation",
        "title": "Apply Spray Foam Insulation",
        "category": "insulation_energy",
        "safety_level": "ppe_required",
    },
    "Clogged Washing Machine": {
        "id": "fix-washing-machine-not-draining",
        "canonical_query": "how to fix a washing machine that will not drain",
        "title": "Fix a Washing Machine That Will Not Drain",
        "category": "appliances",
        "safety_level": "electrical_plumbing",
    },
    "Leaky Asphalt Roof": {
        "id": "repair-leaking-asphalt-shingle-roof",
        "canonical_query": "how to repair a leaking asphalt shingle roof",
        "title": "Repair a Leaking Asphalt Shingle Roof",
        "category": "roofing_gutters",
        "safety_level": "high",
    },
    "Cleaning a Dryer Vent": {
        "id": "clean-dryer-vent",
        "canonical_query": "how to clean a dryer vent",
        "title": "Clean a Dryer Vent",
        "category": "hvac",
        "safety_level": "standard",
    },
    "Driveway Repair": {
        "id": "repair-driveway",
        "canonical_query": "how to repair a driveway",
        "title": "Repair a Driveway",
        "category": "landscape_drainage",
        "safety_level": "standard",
    },
    "Insulating Around a Window": {
        "id": "insulate-around-window",
        "canonical_query": "how to insulate around a window",
        "title": "Insulate Around a Window",
        "category": "doors_windows",
        "safety_level": "standard",
    },
    "Painting a Wood Deck": {
        "id": "paint-wood-deck",
        "canonical_query": "how to paint a wood deck",
        "title": "Paint a Wood Deck",
        "category": "exterior",
        "safety_level": "standard",
    },
    "Whole-House Battery": {
        "id": "install-whole-house-battery",
        "canonical_query": "how to install a whole house battery",
        "title": "Install a Whole-House Battery System",
        "category": "electrical_energy",
        "safety_level": "licensed_or_high_risk",
    },
    "Patio Construction": {
        "id": "build-patio",
        "canonical_query": "how to build a patio",
        "title": "Build a Patio",
        "category": "exterior",
        "safety_level": "standard",
    },
    "Getting Rid of Bats": {
        "id": "remove-bats-from-house",
        "canonical_query": "how to get rid of bats",
        "title": "Remove Bats From a House",
        "category": "pest_control",
        "safety_level": "health_legal",
    },
    "Adding a Panel Circuit": {
        "id": "add-circuit-to-electrical-panel",
        "canonical_query": "how to add a circuit to an electrical panel",
        "title": "Add a Circuit to an Electrical Panel",
        "category": "electrical",
        "safety_level": "licensed_or_high_risk",
    },
    "Installing an RO Filter": {
        "id": "install-reverse-osmosis-water-filter",
        "canonical_query": "how to install a reverse osmosis water filter",
        "title": "Install a Reverse Osmosis Water Filter",
        "category": "plumbing",
        "safety_level": "plumbing",
    },
    "Installing Smart Plugs": {
        "id": "install-smart-plug",
        "canonical_query": "how to install a smart plug",
        "title": "Install a Smart Plug",
        "category": "electrical",
        "safety_level": "low",
    },
    "Patio Repair": {
        "id": "repair-patio",
        "canonical_query": "how to repair a patio",
        "title": "Repair a Patio",
        "category": "exterior",
        "safety_level": "standard",
    },
    "Replacing RO Filters": {
        "id": "replace-reverse-osmosis-filters",
        "canonical_query": "how to replace reverse osmosis filters",
        "title": "Replace Reverse Osmosis Filters",
        "category": "plumbing",
        "safety_level": "plumbing",
    },
    "Electrical Transfer Switch": {
        "id": "install-transfer-switch",
        "canonical_query": "how to install a generator transfer switch",
        "title": "Install a Generator Transfer Switch",
        "category": "electrical",
        "safety_level": "licensed_or_high_risk",
    },
    "Replacing a Storm Window": {
        "id": "replace-storm-window",
        "canonical_query": "how to replace a storm window",
        "title": "Replace a Storm Window",
        "category": "doors_windows",
        "safety_level": "standard",
    },
    "Installing an Aluminum Deck": {
        "id": "install-aluminum-deck",
        "canonical_query": "how to install an aluminum deck",
        "title": "Install an Aluminum Deck",
        "category": "exterior",
        "safety_level": "structural",
    },
    "Vacuuming a Heat Pump": {
        "id": "vacuum-heat-pump-system",
        "canonical_query": "how to vacuum a heat pump system",
        "title": "Vacuum a Heat Pump Refrigerant System",
        "category": "hvac",
        "safety_level": "licensed_or_high_risk",
    },
    "Concrete Anchor Column": {
        "id": "pour-concrete-anchor-column",
        "canonical_query": "how to pour a concrete anchor column",
        "title": "Pour a Concrete Anchor Column",
        "category": "exterior",
        "safety_level": "structural",
    },
}


DROP_CLUSTERS = {
    "Home Improvement": "broad home improvement intent",
    "Home Repair": "broad home repair intent",
    "Home Maintenance": "broad maintenance intent",
    "Appliance Installation": "generic appliance category, no specific install target",
}


DROP_PATTERNS = [
    (re.compile(r"\bnear me\b"), "local service lookup"),
    (re.compile(r"\b(contractor|company|companies|professional|service|installer|technician|repairman|pest control)\b"), "hire-a-pro or service lookup"),
    (re.compile(r"\b(cost|price|estimate|financing|affordable|cheap|expensive)\b"), "cost or buying research"),
    (re.compile(r"\bpermit|requirements?|code requirements?|legal requirements?\b"), "permit, requirement, or code research"),
    (re.compile(r"\bhow long|time required|project timeline\b"), "time estimate rather than walkthrough"),
    (re.compile(r"\btools needed|materials needed|supply list|checklist\b"), "prep list rather than walkthrough"),
    (re.compile(r"\bmistakes|tips|ideas|design|inspiration|best products?\b"), "advice or inspiration rather than walkthrough"),
    (re.compile(r"\bhow often|when to|signs of|common causes?|prevent\b"), "diagnostic or prevention intent"),
]


ACTION_TERMS = re.compile(
    r"\b(how to|diy|step by step|install|replace|repair|fix|clean|paint|build|"
    r"remove|get rid|shingle|tile|insulate|add|use|vacuum|pour|seal|"
    r"can i|best way to|home remedies)\b"
)


def normalize_space(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def slug_text(value):
    clean = normalize_space(value).lower().replace("&", " and ")
    clean = re.sub(r"[^a-z0-9]+", " ", clean)
    return re.sub(r"\s+", " ", clean).strip()


def classify(row):
    phrase = row["search_phrase"]
    cluster = row["topic_cluster"]
    phrase_l = slug_text(phrase)

    if cluster in DROP_CLUSTERS:
        return False, DROP_CLUSTERS[cluster]

    for pattern, reason in DROP_PATTERNS:
        if pattern.search(phrase_l):
            return False, reason

    if cluster not in CANONICAL_BY_CLUSTER:
        return False, "not mapped to a canonical walkthrough yet"

    if not ACTION_TERMS.search(phrase_l):
        return False, "not phrased as an actionable walkthrough"

    return True, "actionable walkthrough variant"


def read_source_rows():
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook["500 Search Phrases"]
    headers = [normalize_space(cell.value) for cell in sheet[4]]
    rows = []

    for excel_row in range(5, sheet.max_row + 1):
        values = [sheet.cell(excel_row, col).value for col in range(1, len(headers) + 1)]
        if not any(values):
            continue
        row = dict(zip(headers, values))
        rows.append(
            {
                "excel_row": excel_row,
                "rank": row.get("Rank"),
                "search_phrase": normalize_space(row.get("Search phrase")),
                "topic_cluster": normalize_space(row.get("Topic cluster")),
                "priority_score": row.get("Priority score"),
                "search_intent": normalize_space(row.get("Search intent")),
                "funnel_stage": normalize_space(row.get("Funnel stage")),
                "recommended_content": normalize_space(row.get("Recommended content")),
                "volume_status": normalize_space(row.get("Volume status")),
                "geography": normalize_space(row.get("Geography")),
                "method_source": normalize_space(row.get("Method source")),
            }
        )

    return rows


def build_outputs():
    rows = read_source_rows()
    kept = []
    removed = []
    by_id = defaultdict(list)

    for row in rows:
        keep, reason = classify(row)
        cluster = row["topic_cluster"]
        if keep:
            canonical = CANONICAL_BY_CLUSTER[cluster]
            record = {
                "canonical_id": canonical["id"],
                "canonical_query": canonical["canonical_query"],
                "canonical_title": canonical["title"],
                "category": canonical["category"],
                "safety_level": canonical["safety_level"],
                "query_variant": row["search_phrase"],
                "normalized_variant": slug_text(row["search_phrase"]),
                "original_rank": row["rank"],
                "source_cluster": cluster,
                "priority_score": row["priority_score"],
                "search_intent": row["search_intent"],
                "review_status": "draft",
            }
            kept.append(record)
            by_id[canonical["id"]].append(record)
        else:
            removed.append(
                {
                    "search_phrase": row["search_phrase"],
                    "original_rank": row["rank"],
                    "source_cluster": cluster,
                    "drop_reason": reason,
                    "priority_score": row["priority_score"],
                    "search_intent": row["search_intent"],
                }
            )

    taxonomy_entries = []
    removed_counter = Counter(item["source_cluster"] for item in removed)
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    for canonical_id in sorted(by_id):
        variants = by_id[canonical_id]
        canonical = next(item for item in CANONICAL_BY_CLUSTER.values() if item["id"] == canonical_id)
        aliases = sorted({canonical["canonical_query"], *[item["query_variant"] for item in variants]})
        taxonomy_entries.append(
            {
                "walkthrough_id": canonical_id,
                "canonical_query": canonical["canonical_query"],
                "title": canonical["title"],
                "category": canonical["category"],
                "safety_level": canonical["safety_level"],
                "review_status": "draft",
                "alias_count": len(aliases),
                "aliases": aliases,
                "source_clusters": sorted({item["source_cluster"] for item in variants}),
                "source_rank_min": min(int(item["original_rank"] or 0) for item in variants),
                "kept_phrase_count": len(variants),
                "removed_phrase_count_from_cluster": int(removed_counter[variants[0]["source_cluster"]]),
            }
        )

    data = {
        "summary": {
            "source_file": str(SOURCE),
            "generated_at": now,
            "source_phrase_count": len(rows),
            "kept_actionable_phrase_count": len(kept),
            "removed_phrase_count": len(removed),
            "canonical_walkthrough_count": len(taxonomy_entries),
            "rule": "Keep only phrases that directly trigger an actionable how-to walkthrough; group variants under one canonical walkthrough ID.",
        },
        "kept_phrases": kept,
        "taxonomy_entries": taxonomy_entries,
        "removed_phrases": removed,
        "drop_reason_counts": dict(Counter(item["drop_reason"] for item in removed)),
    }

    taxonomy = {
        "schema_version": 1,
        "updated_at": now,
        "source": {
            "file": SOURCE.name,
            "edited_by_user": True,
            "rule": data["summary"]["rule"],
        },
        "walkthroughs": {
            item["walkthrough_id"]: {
                "walkthrough_id": item["walkthrough_id"],
                "canonical_query": item["canonical_query"],
                "title": item["title"],
                "category": item["category"],
                "safety_level": item["safety_level"],
                "review_status": item["review_status"],
                "aliases": item["aliases"],
                "source_clusters": item["source_clusters"],
                "kept_phrase_count": item["kept_phrase_count"],
                "removed_phrase_count_from_cluster": item["removed_phrase_count_from_cluster"],
            }
            for item in taxonomy_entries
        },
        "removed_phrase_examples": removed[:120],
        "drop_reason_counts": data["drop_reason_counts"],
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    WORKBOOK_DATA_PATH.write_text(json.dumps(data, indent=2), encoding="utf-8")
    TAXONOMY_PATH.write_text(json.dumps(taxonomy, indent=2), encoding="utf-8")
    return data


if __name__ == "__main__":
    result = build_outputs()
    print(json.dumps(result["summary"], indent=2))
